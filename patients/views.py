from django.shortcuts import render,redirect
from .models import Patient
from .forms import PatientForm,DoctorForm
from django.shortcuts import get_object_or_404
import random
from django.db.models import Q
from .models import Doctor
from .models import Appointment
from .forms import AppointmentForm
from .models import Billing
from .forms import BillingForm
from .models import Medicine
from .forms import MedicineForm
from .models import Laboratory
from .forms import LaboratoryForm,PaymentForm
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from reportlab.pdfgen import canvas
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
from openpyxl import Workbook
from django.contrib.auth import update_session_auth_hash
from .models import ActivityLog,Payment



def home(request):
    return redirect('login')
def patient_list(request):
    search=request.GET.get('search')
    if search:
        patients=Patient.objects.filter(Q (patient_id__icontains=search)|Q(name__icontains=search))
    else:
        patients=Patient.objects.all()
    return render(request,'patients/patient_list.html',{'patients':patients,'search':search})
def add_patient(request):
    if request.method=="POST":
        form=PatientForm(request.POST)
        if form.is_valid():
            patient=form.save(commit=False)
            while True:
                random_id="S"+str(random.randint(10000,99999))
                if Patient.objects.filter(patient_id=random_id).exists():
                    continue
                else:
                    patient.patient_id=random_id
                    break
            patient.save()
            return redirect('patient_list')
    else:
        form=PatientForm()
    return render(request,'patients/add_patient.html',{'form':form})
def edit_patient(request,id):
    patient=get_object_or_404(Patient,id=id)
    if request.method=="POST":
        form=PatientForm(request.POST,instance=patient)
        if form.is_valid():
            form.save()
            return redirect('patient_list')
    else:
        form=PatientForm(instance=patient)
    return render(request,'patients/add_patient.html',{'form':form})
def delete_patient(request,id):
    patient=get_object_or_404(Patient,id=id)
    patient.delete()
    return redirect('patient_list')
def doctor_list(request):
    doctors=Doctor.objects.all()
    return render(request,'patients/doctor_list.html',{'doctors':doctors})


def doctor_list(request):
    search = request.GET.get('search')

    if search:
        doctors = Doctor.objects.filter(
            Q(doctor_id__icontains=search) |
            Q(name__icontains=search)
        )
    else:
        doctors = Doctor.objects.all()

    return render(request, 'patients/doctor_list.html', {
        'doctors': doctors,
        'search': search
    })


def add_doctor(request):
    if request.method == "POST":
        form = DoctorForm(request.POST)
        if form.is_valid():
            doctor = form.save(commit=False)

            while True:
                random_id = "D" + str(random.randint(10000, 99999))
                if Doctor.objects.filter(doctor_id=random_id).exists():
                    continue
                doctor.doctor_id = random_id
                break

            doctor.save()
            return redirect('doctor_list')
    else:
        form = DoctorForm()

    return render(request, 'patients/add_doctor.html', {'form': form})


def edit_doctor(request, id):
    doctor = get_object_or_404(Doctor, id=id)

    if request.method == "POST":
        form = DoctorForm(request.POST, instance=doctor)
        if form.is_valid():
            form.save()
            return redirect('doctor_list')
    else:
        form = DoctorForm(instance=doctor)

    return render(request, 'patients/add_doctor.html', {'form': form})


def delete_doctor(request, id):
    doctor = get_object_or_404(Doctor, id=id)
    doctor.delete()
    return redirect('doctor_list')


def appointment_list(request):
    search = request.GET.get('search')

    if search:
        appointments = Appointment.objects.filter(
            Q(appointment_id__icontains=search) |
            Q(patient__name__icontains=search) |
            Q(doctor__name__icontains=search)
        )
    else:
        appointments = Appointment.objects.all()

    return render(request, 'patients/appointment_list.html', {
        'appointments': appointments,
        'search':search
    })


def add_appointment(request):
    if request.method == "POST":
        form = AppointmentForm(request.POST)

        if form.is_valid():
            appointment = form.save(commit=False)

            while True:
                random_id = "A" + str(random.randint(10000, 99999))
                if not Appointment.objects.filter(appointment_id=random_id).exists():
                    appointment.appointment_id = random_id
                    break

            appointment.save()
            return redirect('appointment_list')
        else:
            print(form.errors)   # <-- This is important

    else:
        form = AppointmentForm()

    return render(request, 'patients/add_appointment.html', {'form': form})


def edit_appointment(request, id):
    appointment = get_object_or_404(Appointment, id=id)

    if request.method == "POST":
        form = AppointmentForm(request.POST, instance=appointment)

        if form.is_valid():
            form.save()
            return redirect('appointment_list')

    else:
        form = AppointmentForm(instance=appointment)

    return render(request, 'patients/add_appointment.html', {'form': form})


def delete_appointment(request, id):
    appointment = get_object_or_404(Appointment, id=id)
    appointment.delete()
    return redirect('appointment_list')



def billing_list(request):
    search = request.GET.get('search')

    if search:
        bills = Billing.objects.filter(
            Q(bill_id__icontains=search) |
            Q(patient__name__icontains=search)
        )
    else:
        bills = Billing.objects.all()

    return render(request, 'patients/billing_list.html', {
        'bills': bills,
        'search': search
    })


def add_billing(request):
    if request.method == "POST":
        form = BillingForm(request.POST)

        if form.is_valid():
            bill = form.save(commit=False)

            while True:
                random_id = "B" + str(random.randint(10000, 99999))
                if not Billing.objects.filter(bill_id=random_id).exists():
                    bill.bill_id = random_id
                    break

            bill.save()
            return redirect('billing_list')

    else:
        form = BillingForm()

    return render(request, 'patients/add_billing.html', {'form': form})


def edit_billing(request, id):
    bill = get_object_or_404(Billing, id=id)

    if request.method == "POST":
        form = BillingForm(request.POST, instance=bill)

        if form.is_valid():
            form.save()
            return redirect('billing_list')

    else:
        form = BillingForm(instance=bill)

    return render(request, 'patients/add_billing.html', {'form': form})


def delete_billing(request, id):
    bill = get_object_or_404(Billing, id=id)
    bill.delete()
    return redirect('billing_list')


def medicine_list(request):
    search = request.GET.get('search')

    if search:
        medicines = Medicine.objects.filter(
            Q(medicine_id__icontains=search) |
            Q(medicine_name__icontains=search)
        )
    else:
        medicines = Medicine.objects.all()

    return render(request, 'patients/medicine_list.html', {
        'medicines': medicines,
        'search': search
    })


def add_medicine(request):
    if request.method == "POST":
        form = MedicineForm(request.POST)

        if form.is_valid():
            medicine = form.save(commit=False)

            while True:
                random_id = "M" + str(random.randint(10000, 99999))
                if not Medicine.objects.filter(medicine_id=random_id).exists():
                    medicine.medicine_id = random_id
                    break

            medicine.save()
            return redirect('medicine_list')

    else:
        form = MedicineForm()

    return render(request, 'patients/add_medicine.html', {'form': form})


def edit_medicine(request, id):
    medicine = get_object_or_404(Medicine, id=id)

    if request.method == "POST":
        form = MedicineForm(request.POST, instance=medicine)

        if form.is_valid():
            form.save()
            return redirect('medicine_list')

    else:
        form = MedicineForm(instance=medicine)

    return render(request, 'patients/add_medicine.html', {'form': form})


def delete_medicine(request, id):
    medicine = get_object_or_404(Medicine, id=id)
    medicine.delete()
    return redirect('medicine_list')


def laboratory_list(request):
    search = request.GET.get('search')

    if search:
        tests = Laboratory.objects.filter(
            Q(test_id__icontains=search) |
            Q(patient__name__icontains=search) |
            Q(test_name__icontains=search)
        )
    else:
        tests = Laboratory.objects.all()

    return render(request, 'patients/laboratory_list.html', {
        'tests': tests,
        'search': search
    })


def add_laboratory(request):
    if request.method == "POST":
        form = LaboratoryForm(request.POST)

        if form.is_valid():
            test = form.save(commit=False)

            while True:
                random_id = "T" + str(random.randint(10000, 99999))
                if not Laboratory.objects.filter(test_id=random_id).exists():
                    test.test_id = random_id
                    break

            test.save()
            return redirect('laboratory_list')

    else:
        form = LaboratoryForm()

    return render(request, 'patients/add_laboratory.html', {'form': form})


def edit_laboratory(request, id):
    test = get_object_or_404(Laboratory, id=id)

    if request.method == "POST":
        form = LaboratoryForm(request.POST, instance=test)

        if form.is_valid():
            form.save()
            return redirect('laboratory_list')

    else:
        form = LaboratoryForm(instance=test)

    return render(request, 'patients/add_laboratory.html', {'form': form})


def delete_laboratory(request, id):
    test = get_object_or_404(Laboratory, id=id)
    test.delete()
    return redirect('laboratory_list')

def login_view(request):
    if request.method == "POST":
        username = request.POST.get("username")
        password = request.POST.get("password")

        user = authenticate(request, username=username, password=password)

        if user is not None:
            login(request, user)
            
            return redirect("dashboard")
        else:
            messages.error(request, "Invalid Username or Password")

    return render(request, "patients/login.html")

def logout_view(request):
    logout(request)
    return redirect("login")
@login_required
def dashboard(request):

    appointments = Appointment.objects.all()

    events = []

    for appointment in appointments:
        events.append({
            "title": f"{appointment.patient.name} - {appointment.doctor.name}",
            "start": appointment.appointment_date.strftime("%Y-%m-%d"),
        })

    context = {
        "total_patients": Patient.objects.count(),
        "total_doctors": Doctor.objects.count(),
        "total_appointments": Appointment.objects.count(),
        "total_bills": Billing.objects.count(),
        "total_medicines": Medicine.objects.count(),
        "total_tests": Laboratory.objects.count(),
        "events": events,
    }

    return render(request, "patients/dashboard.html", context)

def patient_pdf(request):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="patients_report.pdf"'

    p = canvas.Canvas(response)

    p.setFont("Helvetica-Bold", 18)
    p.drawString(180, 800, "Hospital Management System")

    p.setFont("Helvetica", 14)
    p.drawString(230, 775, "Patient Report")

    y = 740

    p.setFont("Helvetica-Bold", 10)
    p.drawString(40, y, "Patient ID")
    p.drawString(120, y, "Name")
    p.drawString(240, y, "Age")
    p.drawString(290, y, "Gender")
    p.drawString(360, y, "Phone")
    p.drawString(470, y, "Disease")

    y -= 20

    p.setFont("Helvetica", 10)

    for patient in Patient.objects.all():
        p.drawString(40, y, patient.patient_id)
        p.drawString(120, y, patient.name)
        p.drawString(240, y, str(patient.age))
        p.drawString(290, y, patient.gender)
        p.drawString(360, y, patient.phone)
        p.drawString(470, y, patient.disease)

        y -= 20

        if y < 40:
            p.showPage()
            y = 800

    p.save()
    return response

def doctor_pdf(request):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="doctors_report.pdf"'

    p = canvas.Canvas(response)

    p.setFont("Helvetica-Bold", 18)
    p.drawString(180, 800, "Hospital Management System")

    p.setFont("Helvetica", 14)
    p.drawString(235, 775, "Doctor Report")

    y = 740

    p.setFont("Helvetica-Bold", 10)
    p.drawString(40, y, "Doctor ID")
    p.drawString(120, y, "Name")
    p.drawString(240, y, "Specialization")
    p.drawString(390, y, "Phone")
    p.drawString(490, y, "Email")

    y -= 20

    p.setFont("Helvetica", 10)

    for doctor in Doctor.objects.all():
        p.drawString(40, y, doctor.doctor_id)
        p.drawString(120, y, doctor.name)
        p.drawString(240, y, doctor.specialization)
        p.drawString(390, y, doctor.phone)
        p.drawString(490, y, str(doctor.email))

        y -= 20

        if y < 40:
            p.showPage()
            y = 800

    p.save()
    return response

def appointment_pdf(request):

    appointments = Appointment.objects.all()

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="appointments.pdf"'

    pdf = SimpleDocTemplate(response)

    elements = []

    styles = getSampleStyleSheet()

    title = Paragraph("Hospital Appointment Report", styles['Title'])
    elements.append(title)

    data = [
        [
            "ID",
            "Patient",
            "Doctor",
            "Date",
            "Time",
            "Status"
        ]
    ]

    for appointment in appointments:
        data.append([
            appointment.appointment_id,
            appointment.patient.name,
            appointment.doctor.name,
            appointment.appointment_date,
            appointment.appointment_time,
            appointment.status
        ])

    table = Table(data)

    table.setStyle(TableStyle([
        ('BACKGROUND',(0,0),(-1,0),'grey'),
        ('GRID',(0,0),(-1,-1),1,'black')
    ]))

    elements.append(table)

    pdf.build(elements)

    return response

def billing_pdf(request):

    bills = Billing.objects.all()

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="billing_report.pdf"'

    pdf = SimpleDocTemplate(response)

    elements = []

    styles = getSampleStyleSheet()

    title = Paragraph("Hospital Billing Report", styles['Title'])
    elements.append(title)

    data = [
        [
            "Bill ID",
            "Patient",
            "Doctor",
            "Consultation",
            "Medicine",
            "Test",
            "Total",
            "Status"
        ]
    ]

    for bill in bills:
        data.append([
            bill.bill_id,
            bill.patient.name,
            bill.doctor.name,
            bill.consultation_fee,
            bill.medicine_fee,
            bill.test_fee,
            bill.total_amount,
            bill.payment_status
        ])

    table = Table(data)

    table.setStyle(TableStyle([
        ('BACKGROUND',(0,0),(-1,0),'grey'),
        ('GRID',(0,0),(-1,-1),1,'black')
    ]))

    elements.append(table)

    pdf.build(elements)

    return response
def medicine_pdf(request):

    medicines = Medicine.objects.all()

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="medicine_report.pdf"'

    pdf = SimpleDocTemplate(response)

    elements = []

    styles = getSampleStyleSheet()

    title = Paragraph("Hospital Medicine Report", styles['Title'])
    elements.append(title)

    data = [
        [
            "Medicine ID",
            "Name",
            "Company",
            "Price",
            "Quantity",
            "Expiry Date"
        ]
    ]

    for medicine in medicines:
        data.append([
            medicine.medicine_id,
            medicine.medicine_name,
            medicine.company,
            medicine.price,
            medicine.quantity,
            medicine.expiry_date
        ])

    table = Table(data)

    table.setStyle(TableStyle([
        ('BACKGROUND',(0,0),(-1,0),'grey'),
        ('GRID',(0,0),(-1,-1),1,'black')
    ]))

    elements.append(table)

    pdf.build(elements)

    return response

def laboratory_pdf(request):

    tests = Laboratory.objects.all()

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="laboratory_report.pdf"'

    pdf = SimpleDocTemplate(response)

    elements = []

    styles = getSampleStyleSheet()

    title = Paragraph("Hospital Laboratory Report", styles['Title'])
    elements.append(title)

    data = [
        [
            "Test ID",
            "Patient",
            "Test Name",
            "Test Date",
            "Result",
            "Amount"
        ]
    ]

    for test in tests:
        data.append([
            test.test_id,
            test.patient.name,
            test.test_name,
            test.test_date,
            test.result,
            test.amount
        ])

    table = Table(data)

    table.setStyle(TableStyle([
        ('BACKGROUND',(0,0),(-1,0),'grey'),
        ('GRID',(0,0),(-1,-1),1,'black')
    ]))

    elements.append(table)

    pdf.build(elements)

    return response


def patient_excel(request):

    patients = Patient.objects.all()

    wb = Workbook()

    ws = wb.active
    ws.title = "Patients"

    ws.append([
        "Patient ID",
        "Name",
        "Age",
        "Gender",
        "Phone",
        "Disease"
    ])

    for patient in patients:
        ws.append([
            patient.patient_id,
            patient.name,
            patient.age,
            patient.gender,
            patient.phone,
            patient.disease
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response['Content-Disposition'] = 'attachment; filename=patients.xlsx'

    wb.save(response)

    return response

def doctor_excel(request):

    doctors = Doctor.objects.all()

    wb = Workbook()

    ws = wb.active
    ws.title = "Doctors"

    ws.append([
        "Doctor ID",
        "Name",
        "Specialization",
        "Phone",
        "Email"
    ])

    for doctor in doctors:
        ws.append([
            doctor.doctor_id,
            doctor.name,
            doctor.specialization,
            doctor.phone,
            doctor.email
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response['Content-Disposition'] = 'attachment; filename=doctors.xlsx'

    wb.save(response)

    return response


def appointment_excel(request):

    appointments = Appointment.objects.all()

    wb = Workbook()

    ws = wb.active
    ws.title = "Appointments"

    ws.append([
        "Appointment ID",
        "Patient",
        "Doctor",
        "Date",
        "Time",
        "Status"
    ])

    for appointment in appointments:
        ws.append([
            appointment.appointment_id,
            appointment.patient.name,
            appointment.doctor.name,
            appointment.appointment_date,
            appointment.appointment_time,
            appointment.status
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response['Content-Disposition'] = 'attachment; filename=appointments.xlsx'

    wb.save(response)

    return response


def billing_excel(request):

    bills = Billing.objects.all()

    wb = Workbook()

    ws = wb.active
    ws.title = "Billing"

    ws.append([
        "Bill ID",
        "Patient",
        "Doctor",
        "Consultation Fee",
        "Medicine Fee",
        "Test Fee",
        "Total Amount",
        "Payment Status"
    ])

    for bill in bills:
        ws.append([
            bill.bill_id,
            bill.patient.name,
            bill.doctor.name,
            bill.consultation_fee,
            bill.medicine_fee,
            bill.test_fee,
            bill.total_amount,
            bill.payment_status
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response['Content-Disposition'] = 'attachment; filename=billing.xlsx'

    wb.save(response)

    return response

def medicine_excel(request):

    medicines = Medicine.objects.all()

    wb = Workbook()

    ws = wb.active
    ws.title = "Medicines"

    ws.append([
        "Medicine ID",
        "Medicine Name",
        "Company",
        "Price",
        "Quantity",
        "Expiry Date"
    ])

    for medicine in medicines:
        ws.append([
            medicine.medicine_id,
            medicine.medicine_name,
            medicine.company,
            medicine.price,
            medicine.quantity,
            medicine.expiry_date
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response['Content-Disposition'] = 'attachment; filename=medicines.xlsx'

    wb.save(response)

    return response


def laboratory_excel(request):

    tests = Laboratory.objects.all()

    wb = Workbook()

    ws = wb.active
    ws.title = "Laboratory"

    ws.append([
        "Test ID",
        "Patient",
        "Test Name",
        "Test Date",
        "Result",
        "Amount"
    ])

    for test in tests:
        ws.append([
            test.test_id,
            test.patient.name,
            test.test_name,
            test.test_date,
            test.result,
            test.amount
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response['Content-Disposition'] = 'attachment; filename=laboratory.xlsx'

    wb.save(response)

    return response



@login_required
def profile(request):

    return render(request, 'patients/profile.html')


@login_required
def change_password(request):
    if request.method == "POST":
        old_password = request.POST.get("old_password")
        new_password = request.POST.get("new_password")
        confirm_password=request.POST.get("confirm_password")
        user = request.user

        if not user.check_password(old_password):
            messages.error(request,"old password is incorrect")
        elif new_password !=confirm_password:
            messages.error(request,"new password do not match")
        else:
            user.set_password(new_password)
            user.save()

            update_session_auth_hash(request, user)

            messages.success(
                request,
                "Password changed successfully"
            )

            return redirect('profile')

    return render(request, 'patients/change_password.html')



@login_required
def activity_log(request):
    logs = ActivityLog.objects.all().order_by('-action_time')
    return render(request, "patients/activity_log.html", {"logs": logs})


@login_required
def settings(request):
    return render(request, "patients/settings.html")

@login_required
def about(request):
    return render(request, "patients/about.html")


@login_required
def add_payment(request):

    if request.method == "POST":

        form = PaymentForm(request.POST)

        if form.is_valid():
            form.save()

            messages.success(
                request,
                "Payment added successfully."
            )

            return redirect('payment_list')

    else:
        form = PaymentForm()

    return render(
        request,
        'patients/add_payment.html',
        {'form': form}
    )

@login_required
def payment_list(request):

    search = request.GET.get('search')

    if search:
        payments = Payment.objects.filter(
            Q(transaction_id__icontains=search) |
            Q(bill__bill_id__icontains=search)
        ).order_by('-payment_date')
    else:
        payments = Payment.objects.all().order_by('-payment_date')

    return render(
        request,
        'patients/payment_list.html',
        {
            'payments': payments,
            'search': search
        }
    )

@login_required
def payment_pdf(request):

    payments = Payment.objects.all().order_by('-payment_date')

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="payment_report.pdf"'

    pdf = SimpleDocTemplate(response)

    elements = []

    styles = getSampleStyleSheet()

    title = Paragraph(
        "Hospital Payment Report",
        styles['Title']
    )

    elements.append(title)

    data = [
        [
            "Bill ID",
            "Amount",
            "Transaction ID",
            "Status",
            "Payment Date"
        ]
    ]

    for payment in payments:
        data.append([
            payment.bill.bill_id,
            payment.amount,
            payment.transaction_id,
            payment.status,
            payment.payment_date
        ])

    table = Table(data)

    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), 'grey'),
        ('TEXTCOLOR', (0, 0), (-1, 0), 'white'),
        ('GRID', (0, 0), (-1, -1), 1, 'black'),
        ('ALIGN', (1, 1), (1, -1), 'RIGHT'),
    ]))

    elements.append(table)

    pdf.build(elements)

    return response
# Create your views here.

